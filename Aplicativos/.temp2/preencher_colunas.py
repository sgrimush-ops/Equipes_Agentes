"""
Script para preencher as colunas vazias do itens_extraidos.xlsx.

ETAPA 1: Usa o arquivo_codigos_transicao.txt para cruzar Código Transição
         e descobrir o CODIGO_PRODUTO (= Codigo Consico) e DESCRICAO_PRODUTO (= Descrição Consico).

ETAPA 2: Gera uma query SQL para o usuário extrair do Consinco os dados
         de fornecedor (Código e Descrição) usando os códigos de produto já descobertos.
         Após a extração, o script faz o merge final.
"""
import os
from pathlib import Path
import pandas as pd

if __name__ == '__main__':
    os.chdir(Path(__file__).parent.resolve())


def carregar_mapa_transicao() -> pd.DataFrame:
    """Carrega o arquivo de códigos de transição e retorna mapa CODIGO_TRANSICAO -> CODIGO_PRODUTO + DESCRICAO."""
    caminho = Path(__file__).parent.parent / "import_querys" / "arquivo_codigos_transicao.txt"
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo de transição não encontrado: {caminho}")

    df = pd.read_csv(caminho, sep=';', encoding='latin-1', dtype=str)
    print(f"[MAPA] Carregado: {len(df)} linhas, {df['CODIGO_TRANSICAO'].notna().sum()} com transição")

    df = df[df['CODIGO_TRANSICAO'].notna()].copy()
    df['CODIGO_TRANSICAO'] = df['CODIGO_TRANSICAO'].str.strip()
    df = df[df['CODIGO_TRANSICAO'] != '']

    mapa = df.drop_duplicates(subset='CODIGO_TRANSICAO', keep='first')[
        ['CODIGO_TRANSICAO', 'CODIGO_PRODUTO', 'DESCRICAO_PRODUTO']
    ].copy()

    print(f"[MAPA] Mapa único: {len(mapa)} códigos de transição distintos")
    return mapa


def etapa1_preencher_consico(mapa: pd.DataFrame) -> pd.DataFrame:
    """Preenche Codigo Consico e Descrição Consico via cruzamento com código de transição."""
    caminho_excel = Path("itens_extraidos.xlsx")
    if not caminho_excel.exists():
        raise FileNotFoundError(f"Excel não encontrado: {caminho_excel}")

    print(f"\n[ETAPA 1] Lendo {caminho_excel}...")
    df = pd.read_excel(caminho_excel, engine='openpyxl', dtype=str)
    print(f"[ETAPA 1] {len(df)} linhas carregadas")

    colunas = list(df.columns)
    col_transicao = colunas[0]
    col_cod_consico = colunas[8]
    col_desc_consico = colunas[9]

    print(f"[ETAPA 1] Coluna transição: '{col_transicao}'")
    print(f"[ETAPA 1] Coluna cod consico: '{col_cod_consico}'")
    print(f"[ETAPA 1] Coluna desc consico: '{col_desc_consico}'")

    df['_chave'] = df[col_transicao].astype(str).str.strip()
    df['_chave'] = df['_chave'].str.replace(r'\.0$', '', regex=True)

    mapa['_chave'] = mapa['CODIGO_TRANSICAO'].astype(str).str.strip()
    mapa['_chave_sem_zero'] = mapa['_chave'].str.lstrip('0')

    merge_map = mapa.set_index('_chave')[['CODIGO_PRODUTO', 'DESCRICAO_PRODUTO']].to_dict('index')
    mapa_dedup = mapa.drop_duplicates(subset='_chave_sem_zero', keep='first')
    merge_map_sem_zero = mapa_dedup.set_index('_chave_sem_zero')[['CODIGO_PRODUTO', 'DESCRICAO_PRODUTO']].to_dict('index')

    preenchidos = 0
    nao_encontrados = set()

    for idx in df.index:
        chave = df.at[idx, '_chave']
        if chave in ('nan', 'None', ''):
            continue

        dados = merge_map.get(chave)
        if not dados:
            chave_limpa = str(chave).lstrip('0')
            dados = merge_map_sem_zero.get(chave_limpa)

        if dados:
            df.at[idx, col_cod_consico] = str(dados['CODIGO_PRODUTO'])
            df.at[idx, col_desc_consico] = str(dados['DESCRICAO_PRODUTO'])
            preenchidos += 1
        else:
            nao_encontrados.add(chave)

    df.drop(columns=['_chave'], inplace=True)

    print(f"\n[ETAPA 1 - RESULTADO]")
    print(f"  Linhas preenchidas (Codigo/Descrição Consico): {preenchidos:,} / {len(df):,}")
    print(f"  Códigos de transição não encontrados: {len(nao_encontrados):,}")

    if nao_encontrados and len(nao_encontrados) <= 30:
        print(f"  Códigos faltantes: {sorted(str(x) for x in nao_encontrados)}")
    elif nao_encontrados:
        amostra = sorted(str(x) for x in nao_encontrados)[:20]
        print(f"  Amostra de códigos faltantes (20 de {len(nao_encontrados)}): {amostra}")

    return df


def etapa2_gerar_query_fornecedor(df: pd.DataFrame) -> None:
    """Gera query SQL para extrair CODIGO_FORNECEDOR e DESCRICAO_FORNECEDOR dos produtos encontrados.

    Oracle limita IN() a 1000 expressões (ORA-01795). Divide em blocos de 999 com OR.
    """
    colunas = list(df.columns)
    col_cod_consico = colunas[8]

    codigos = df[col_cod_consico].dropna().unique()
    codigos_int = []
    for c in codigos:
        try:
            codigos_int.append(int(float(c)))
        except (ValueError, TypeError):
            pass

    codigos_int = sorted(set(codigos_int))
    print(f"\n[ETAPA 2] {len(codigos_int)} códigos de produto únicos encontrados para buscar fornecedor")

    CHUNK_SIZE = 999
    blocos = [codigos_int[i:i + CHUNK_SIZE] for i in range(0, len(codigos_int), CHUNK_SIZE)]
    print(f"[ETAPA 2] Dividido em {len(blocos)} blocos de até {CHUNK_SIZE} códigos (anti ORA-01795)")

    filtros_in = []
    for bloco in blocos:
        lista = ', '.join(str(c) for c in bloco)
        filtros_in.append(f"P.SEQPRODUTO IN ({lista})")

    where_clause = '\n    OR '.join(filtros_in)

    query = f"""SELECT
    P.SEQPRODUTO AS CODIGO_PRODUTO,
    P.DESCCOMPLETA AS DESCRICAO_PRODUTO,
    NVL(FORN.CODIGO_FORNECEDOR, 0) AS CODIGO_FORNECEDOR,
    NVL(FORN.FORNECEDOR, 'SEM FORNECEDOR') AS DESCRICAO_FORNECEDOR
FROM MAP_PRODUTO P
LEFT JOIN (
    SELECT
        F.SEQFAMILIA,
        MAX(F.SEQFORNECEDOR) AS CODIGO_FORNECEDOR,
        MAX(G.NOMERAZAO) AS FORNECEDOR
    FROM MAP_FAMFORNEC F
    INNER JOIN GE_PESSOA G
        ON G.SEQPESSOA = F.SEQFORNECEDOR
    WHERE F.PRINCIPAL = 'S'
    GROUP BY F.SEQFAMILIA
) FORN
    ON FORN.SEQFAMILIA = P.SEQFAMILIA
WHERE ({where_clause})
ORDER BY P.SEQPRODUTO"""

    query_path = Path("query_fornecedor.sql")
    with open(query_path, 'w', encoding='utf-8') as f:
        f.write(query)
    print(f"[ETAPA 2] Query SQL salva em: {query_path.resolve()}")
    print(f"[ETAPA 2] Execute essa query no Consinco e salve o resultado como 'fornecedor.txt' (CSV com ;)")
    print(f"[ETAPA 2] Após ter o arquivo, rode: python preencher_colunas.py --merge-fornecedor")

    return df


def etapa3_merge_fornecedor(df: pd.DataFrame) -> pd.DataFrame:
    """Faz o merge dos dados de fornecedor após a extração do Consinco."""
    caminho_forn = Path("fornecedor.txt")
    if not caminho_forn.exists():
        print(f"\n[ETAPA 3] Arquivo '{caminho_forn}' não encontrado.")
        print(f"[ETAPA 3] Execute a query 'query_fornecedor.sql' no Consinco primeiro.")
        return df

    print(f"\n[ETAPA 3] Lendo {caminho_forn}...")

    try:
        df_forn = pd.read_csv(caminho_forn, sep=';', encoding='utf-8-sig', dtype=str)
    except UnicodeDecodeError:
        df_forn = pd.read_csv(caminho_forn, sep=';', encoding='latin-1', dtype=str)

    print(f"[ETAPA 3] {len(df_forn)} registros de fornecedor carregados")
    print(f"[ETAPA 3] Colunas: {list(df_forn.columns)}")

    col_forn_codigo = None
    col_forn_desc = None
    col_forn_prod = None

    for c in df_forn.columns:
        cu = c.upper().strip()
        if 'CODIGO_FORNECEDOR' in cu or 'COD_FORNECEDOR' in cu:
            col_forn_codigo = c
        elif 'DESCRICAO_FORNECEDOR' in cu or 'FORNECEDOR' in cu:
            col_forn_desc = c
        elif 'CODIGO_PRODUTO' in cu or 'SEQPRODUTO' in cu:
            col_forn_prod = c

    if not all([col_forn_codigo, col_forn_desc, col_forn_prod]):
        print(f"[ETAPA 3] ERRO: Não foi possível identificar as colunas automaticamente.")
        print(f"  Esperado: CODIGO_PRODUTO, CODIGO_FORNECEDOR, DESCRICAO_FORNECEDOR (ou FORNECEDOR)")
        return df

    print(f"[ETAPA 3] Mapeando: {col_forn_prod} -> {col_forn_codigo}, {col_forn_desc}")

    df_forn[col_forn_prod] = df_forn[col_forn_prod].astype(str).str.strip()
    forn_map = df_forn.drop_duplicates(subset=col_forn_prod, keep='first').set_index(col_forn_prod)[[col_forn_codigo, col_forn_desc]].to_dict('index')

    colunas = list(df.columns)
    col_cod_consico = colunas[8]
    col_cod_fornecedor = colunas[10]
    col_desc_fornecedor = colunas[11]

    preenchidos = 0
    for idx in df.index:
        cod = df.at[idx, col_cod_consico]
        if pd.isna(cod) or str(cod).strip() in ('', 'nan', 'None'):
            continue
        cod_str = str(cod).strip()

        dados = forn_map.get(cod_str)
        if dados:
            df.at[idx, col_cod_fornecedor] = str(dados[col_forn_codigo])
            df.at[idx, col_desc_fornecedor] = str(dados[col_forn_desc])
            preenchidos += 1

    print(f"[ETAPA 3] Linhas preenchidas com fornecedor: {preenchidos:,}")
    return df


def salvar_resultado(df: pd.DataFrame) -> None:
    """Salva o resultado editando o Excel original via openpyxl para preservar formatação e fórmulas."""
    import shutil
    import openpyxl

    caminho_original = Path("itens_extraidos.xlsx")
    caminho_saida = Path("itens_extraidos_preenchido.xlsx")

    print(f"\n[SALVANDO] Copiando original e preenchendo colunas via openpyxl...")
    shutil.copy2(caminho_original, caminho_saida)

    wb = openpyxl.load_workbook(caminho_saida)
    ws = wb.active

    colunas = list(df.columns)
    col_cod_consico = colunas[8]
    col_desc_consico = colunas[9]
    col_cod_fornecedor = colunas[10]
    col_desc_fornecedor = colunas[11]

    preenchidos = 0
    for i, idx in enumerate(df.index):
        row_excel = i + 2  # linha 1 = header

        val_cod = df.at[idx, col_cod_consico]
        val_desc = df.at[idx, col_desc_consico]
        val_forn_cod = df.at[idx, col_cod_fornecedor]
        val_forn_desc = df.at[idx, col_desc_fornecedor]

        if pd.notna(val_cod) and str(val_cod).strip() not in ('', 'nan', 'None'):
            try:
                ws.cell(row=row_excel, column=9, value=int(float(val_cod)))
            except (ValueError, TypeError):
                ws.cell(row=row_excel, column=9, value=str(val_cod))

        if pd.notna(val_desc) and str(val_desc).strip() not in ('', 'nan', 'None'):
            ws.cell(row=row_excel, column=10, value=str(val_desc))

        if pd.notna(val_forn_cod) and str(val_forn_cod).strip() not in ('', 'nan', 'None'):
            try:
                ws.cell(row=row_excel, column=11, value=int(float(val_forn_cod)))
            except (ValueError, TypeError):
                ws.cell(row=row_excel, column=11, value=str(val_forn_cod))

        if pd.notna(val_forn_desc) and str(val_forn_desc).strip() not in ('', 'nan', 'None'):
            ws.cell(row=row_excel, column=12, value=str(val_forn_desc))
            preenchidos += 1

    wb.save(caminho_saida)
    wb.close()
    print(f"[SALVANDO] {preenchidos:,} linhas escritas, formatação original preservada")
    print(f"[SALVANDO] Concluído! Arquivo: {caminho_saida.resolve()}")


def main() -> None:
    """Função principal."""
    import sys

    print("=" * 60)
    print("  PREENCHIMENTO DE COLUNAS — itens_extraidos.xlsx")
    print("=" * 60)

    merge_mode = '--merge-fornecedor' in sys.argv

    mapa = carregar_mapa_transicao()
    df = etapa1_preencher_consico(mapa)

    if merge_mode:
        df = etapa3_merge_fornecedor(df)
    else:
        etapa2_gerar_query_fornecedor(df)

    salvar_resultado(df)

    print()
    print("=" * 60)
    print("  PROCESSO FINALIZADO")
    print("=" * 60)


if __name__ == '__main__':
    main()
