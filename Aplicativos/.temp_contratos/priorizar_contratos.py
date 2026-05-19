from __future__ import annotations

import argparse
import difflib
import json
import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
PDF_DIR = BASE_DIR / "contratospdf"
OUTPUT_DIR = BASE_DIR / "saida"
DEFAULT_PRIORITY_PATH = BASE_DIR / "relatorio" / "lista_prioridade.xlsx"
CONSOLIDATED_JSON_PATH = OUTPUT_DIR / "contratos_consolidados.json"
MATCHED_PDF_DIR = OUTPUT_DIR / "prioridade_pdfs"
OUTPUT_REPORT_PATH = OUTPUT_DIR / "prioridade_contratos.xlsx"

IGNORED_TOKENS = {
    "LTDA",
    "LTD",
    "S",
    "A",
    "SA",
    "ME",
    "EPP",
    "EIRELI",
    "IND",
    "INDUSTRIA",
    "INDUSTRIAL",
    "COM",
    "COMERCIO",
    "COMERCIAL",
    "DISTRIBUIDORA",
    "DISTRIBUIDOR",
    "DISTR",
    "DIST",
    "DE",
    "DA",
    "DO",
    "E",
    "CIA",
    "ALIMENTOS",
    "BEBIDAS",
    "COSMETICOS",
    "PRODUTO",
    "PRODUTOS",
    "BRASIL",
    "PAPEIS",
    "PLASTICOS",
    "EMBALAGENS",
}


@dataclass
class CandidateMatch:
    status: str
    strategy: str
    match_basis: str
    confidence: float
    pdf_path: Path | None
    contract_record: dict[str, object] | None
    matched_name: str


def normalize_text(value: object) -> str:
    ascii_value = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = ascii_value.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.upper()
    ascii_value = re.sub(r"[^A-Z0-9]+", " ", ascii_value)
    return re.sub(r"\s+", " ", ascii_value).strip()


def normalize_cnpj(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 14 else ""


def build_lookup_key(value: object) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    tokens = [
        token for token in normalized.split()
        if token not in IGNORED_TOKENS and len(token) >= 3
    ]
    return " ".join(tokens)


def sequence_ratio(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    return difflib.SequenceMatcher(None, left, right).ratio()


def token_overlap_score(left: str, right: str) -> float:
    left_tokens = {token for token in left.split() if len(token) >= 3}
    right_tokens = {token for token in right.split() if len(token) >= 3}
    if not left_tokens or not right_tokens:
        return 0.0
    intersection = left_tokens & right_tokens
    union = left_tokens | right_tokens
    if not union:
        return 0.0
    return len(intersection) / len(union)


def combined_score(left: str, right: str) -> float:
    ratio = sequence_ratio(left, right)
    overlap = token_overlap_score(left, right)
    if left == right:
        return 1.0
    if (
        left and right
        and min(len(left), len(right)) >= 5
        and (left in right or right in left)
    ):
        return max(ratio, overlap, 0.95)
    return max(ratio, overlap)


def has_meaningful_overlap(left: str, right: str) -> bool:
    left_tokens = {token for token in left.split() if len(token) >= 3}
    right_tokens = {token for token in right.split() if len(token) >= 3}
    overlap = left_tokens & right_tokens
    return bool(overlap)


def iter_priority_rows(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    cnpj_by_code: dict[str, str] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell or "").strip().lower() for cell in rows[0]]
        code_index = next(
            (
                index for index, value in enumerate(header)
                if value in {"cod", "codigo", "código"}
            ),
            None,
        )
        cnpj_index = next(
            (
                index for index, value in enumerate(header)
                if "cnpj" in value
            ),
            None,
        )
        if code_index is None or cnpj_index is None:
            continue
        for row in rows[1:]:
            code_value = (
                "" if row[code_index] is None
                else str(row[code_index]).strip()
            )
            cnpj_value = normalize_cnpj(row[cnpj_index])
            if code_value and cnpj_value:
                cnpj_by_code[code_value] = cnpj_value

    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell or "").strip().lower() for cell in rows[0]]
    code_index = next(
        (
            index for index, value in enumerate(header)
            if value in {"cod", "codigo", "código"}
        ),
        None,
    )
    supplier_index = next(
        (index for index, value in enumerate(header) if "fornecedor" in value),
        None,
    )
    if code_index is None or supplier_index is None:
        raise ValueError(
            "Nao foi possivel identificar as colunas Cod e "
            "Fornecedor na planilha."
        )

    items: list[dict[str, object]] = []
    for row in rows[1:]:
        supplier_name = str(row[supplier_index] or "").strip()
        if not supplier_name:
            continue
        code_value = row[code_index]
        items.append(
            {
                "codigo_prioridade": (
                    "" if code_value is None else str(code_value).strip()
                ),
                "fornecedor_prioridade": supplier_name,
                "chave_prioridade": build_lookup_key(supplier_name),
                "cnpj_prioridade": cnpj_by_code.get(
                    "" if code_value is None else str(code_value).strip(),
                    "",
                ),
            }
        )
    return items


def load_contract_records() -> list[dict[str, object]]:
    if not CONSOLIDATED_JSON_PATH.exists():
        return []
    payload = json.loads(CONSOLIDATED_JSON_PATH.read_text(encoding="utf-8"))
    records: list[dict[str, object]] = []
    for item in payload:
        record = dict(item)
        supplier_name = (
            record.get("fornecedor_descricao_atual")
            or record.get("fornecedor_nome")
            or ""
        )
        record["_lookup_name"] = str(supplier_name).strip()
        record["_lookup_key"] = build_lookup_key(supplier_name)
        record["_lookup_code"] = str(
            record.get("codigo_fornecedor_consinco", "") or ""
        ).strip()
        record["_lookup_cnpj"] = normalize_cnpj(record.get("cnpj"))
        records.append(record)
    return records


def load_pdf_records() -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for pdf_path in sorted(PDF_DIR.glob("*.pdf")):
        records.append(
            {
                "pdf_path": pdf_path,
                "pdf_name": pdf_path.name,
                "_lookup_name": pdf_path.stem,
                "_lookup_key": build_lookup_key(pdf_path.stem),
            }
        )
    return records


def select_best_match(
    priority_key: str,
    records: list[dict[str, object]],
    status: str,
    min_score: float,
) -> CandidateMatch | None:
    exact_matches = [
        item for item in records if item.get("_lookup_key") == priority_key
    ]
    if exact_matches:
        chosen = exact_matches[0]
        pdf_path = (
            chosen.get("pdf_path")
            if status == "pdf_localizado"
            else PDF_DIR / str(chosen.get("arquivo_pdf", ""))
        )
        return CandidateMatch(
            status=status,
            strategy="exato",
            match_basis="nome",
            confidence=1.0,
            pdf_path=pdf_path,
            contract_record=chosen if status == "contrato_extraido" else None,
            matched_name=str(chosen.get("_lookup_name", "")),
        )

    ranked: list[tuple[float, dict[str, object]]] = []
    for item in records:
        candidate_key = str(item.get("_lookup_key", ""))
        if not has_meaningful_overlap(priority_key, candidate_key):
            continue
        score = combined_score(priority_key, candidate_key)
        if score >= min_score:
            ranked.append((score, item))
    if not ranked:
        return None

    ranked.sort(
        key=lambda pair: (
            pair[0],
            len(str(pair[1].get("_lookup_key", ""))),
        ),
        reverse=True,
    )
    score, chosen = ranked[0]
    pdf_path = (
        chosen.get("pdf_path")
        if status == "pdf_localizado"
        else PDF_DIR / str(chosen.get("arquivo_pdf", ""))
    )
    return CandidateMatch(
        status=status,
        strategy="fuzzy",
        match_basis="nome",
        confidence=round(score, 4),
        pdf_path=pdf_path,
        contract_record=chosen if status == "contrato_extraido" else None,
        matched_name=str(chosen.get("_lookup_name", "")),
    )


def select_contract_match_by_field(
    field_name: str,
    field_value: str,
    records: list[dict[str, object]],
) -> CandidateMatch | None:
    if not field_value:
        return None
    for item in records:
        if str(item.get(field_name, "") or "").strip() != field_value:
            continue
        pdf_path = PDF_DIR / str(item.get("arquivo_pdf", ""))
        basis = "codigo" if field_name == "_lookup_code" else "cnpj"
        return CandidateMatch(
            status="contrato_extraido",
            strategy="exato",
            match_basis=basis,
            confidence=1.0,
            pdf_path=pdf_path,
            contract_record=item,
            matched_name=str(item.get("_lookup_name", "")),
        )
    return None


def find_priority_match(
    priority_item: dict[str, object],
    contract_records: list[dict[str, object]],
    pdf_records: list[dict[str, object]],
) -> CandidateMatch:
    priority_code = str(
        priority_item.get("codigo_prioridade", "") or ""
    ).strip()
    priority_key = str(priority_item.get("chave_prioridade", ""))
    priority_cnpj = normalize_cnpj(priority_item.get("cnpj_prioridade"))

    code_match = select_contract_match_by_field(
        "_lookup_code",
        priority_code,
        contract_records,
    )
    if code_match:
        return code_match

    cnpj_match = select_contract_match_by_field(
        "_lookup_cnpj",
        priority_cnpj,
        contract_records,
    )
    if cnpj_match:
        return cnpj_match

    if not priority_key:
        return CandidateMatch(
            "nao_encontrado",
            "sem_chave",
            "nenhum",
            0.0,
            None,
            None,
            "",
        )

    contract_match = select_best_match(
        priority_key,
        contract_records,
        status="contrato_extraido",
        min_score=0.78,
    )
    if contract_match:
        return contract_match

    pdf_match = select_best_match(
        priority_key,
        pdf_records,
        status="pdf_localizado",
        min_score=0.78,
    )
    if pdf_match:
        return pdf_match

    return CandidateMatch(
        "nao_encontrado",
        "sem_match",
        "nenhum",
        0.0,
        None,
        None,
        "",
    )


def copy_matched_pdf(pdf_path: Path | None) -> str:
    if pdf_path is None or not pdf_path.exists():
        return ""
    MATCHED_PDF_DIR.mkdir(parents=True, exist_ok=True)
    target_path = MATCHED_PDF_DIR / pdf_path.name
    if not target_path.exists():
        shutil.copy2(pdf_path, target_path)
    return str(target_path)


def build_result_rows(
    priority_items: list[dict[str, object]],
    contract_records: list[dict[str, object]],
    pdf_records: list[dict[str, object]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for item in priority_items:
        match = find_priority_match(item, contract_records, pdf_records)
        contract_record = match.contract_record or {}
        copied_pdf = copy_matched_pdf(match.pdf_path)
        rows.append(
            {
                "codigo_prioridade": item.get("codigo_prioridade", ""),
                "fornecedor_prioridade": item.get("fornecedor_prioridade", ""),
                "status_match": match.status,
                "estrategia_match": match.strategy,
                "base_match": match.match_basis,
                "confianca_match": match.confidence,
                "nome_encontrado": match.matched_name,
                "cnpj_prioridade": item.get("cnpj_prioridade", ""),
                "arquivo_pdf": (
                    "" if match.pdf_path is None else match.pdf_path.name
                ),
                "pdf_copiado": copied_pdf,
                "fornecedor_contrato": contract_record.get(
                    "fornecedor_nome",
                    "",
                ),
                "fornecedor_cadastro_atual": contract_record.get(
                    "fornecedor_descricao_atual",
                    "",
                ),
                "codigo_fornecedor_consinco": contract_record.get(
                    "codigo_fornecedor_consinco",
                    "",
                ),
                "cnpj": contract_record.get("cnpj", ""),
                "a_receber": contract_record.get("a_receber", ""),
                "dias_de_atraso": contract_record.get("dias_de_atraso", ""),
                "data_assinatura": contract_record.get("data_assinatura", ""),
                "ano_assinatura": contract_record.get("ano_assinatura", ""),
                "layout_tipo": contract_record.get("layout_tipo", ""),
            }
        )
    return rows


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [
            "" if cell.value is None else str(cell.value)
            for cell in column_cells
        ]
        width = min(max(len(value) for value in values) + 2, 60)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = width


def append_sheet(
    workbook: Workbook,
    title: str,
    rows: list[dict[str, object]],
) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet.append(["sem_dados"])
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    autosize_columns(worksheet)


def summarize_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    status_counter: dict[str, int] = {}
    for row in rows:
        status = str(row.get("status_match", "nao_encontrado"))
        status_counter[status] = status_counter.get(status, 0) + 1

    summary = [
        {"indicador": "total_fornecedores_prioridade", "valor": len(rows)},
        {
            "indicador": "contrato_extraido",
            "valor": status_counter.get("contrato_extraido", 0),
        },
        {
            "indicador": "pdf_localizado",
            "valor": status_counter.get("pdf_localizado", 0),
        },
        {
            "indicador": "nao_encontrado",
            "valor": status_counter.get("nao_encontrado", 0),
        },
    ]
    return summary


def write_report(rows: list[dict[str, object]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ordered_rows = sorted(
        rows,
        key=lambda row: (
            {
                "contrato_extraido": 0,
                "pdf_localizado": 1,
                "nao_encontrado": 2,
            }.get(
                str(row.get("status_match", "nao_encontrado")),
                3,
            ),
            str(row.get("fornecedor_prioridade", "")),
        ),
    )
    append_sheet(workbook, "resumo", summarize_rows(ordered_rows))
    append_sheet(workbook, "prioridades", ordered_rows)
    append_sheet(
        workbook,
        "nao_encontrados",
        [
            row for row in ordered_rows
            if row.get("status_match") == "nao_encontrado"
        ],
    )
    append_sheet(
        workbook,
        "pendentes_cnpj",
        [
            row for row in ordered_rows
            if row.get("status_match") == "nao_encontrado"
            and not row.get("cnpj_prioridade")
        ],
    )
    append_sheet(
        workbook,
        "revisar_fuzzy",
        [
            row for row in ordered_rows
            if row.get("estrategia_match") == "fuzzy"
            and row.get("base_match") == "nome"
        ],
    )
    workbook.save(output_path)


def write_pending_cnpj_sheet(
    source_workbook_path: Path,
    rows: list[dict[str, object]],
) -> None:
    workbook = load_workbook(source_workbook_path)
    sheet_name = "CNPJ_AUXILIAR"
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        workbook.remove(worksheet)

    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(["Cod", "Fornecedor", "CNPJ", "Observacao"])

    pending_rows = [
        row for row in rows
        if row.get("status_match") == "nao_encontrado"
    ]
    pending_rows.sort(
        key=lambda row: str(row.get("fornecedor_prioridade", "")),
    )

    for row in pending_rows:
        worksheet.append(
            [
                row.get("codigo_prioridade", ""),
                row.get("fornecedor_prioridade", ""),
                row.get("cnpj_prioridade", ""),
                "Preencher CNPJ para nova reconciliacao automatica",
            ]
        )

    autosize_columns(worksheet)
    workbook.save(source_workbook_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Cruza a lista de prioridade com contratos extraidos e PDFs "
            "locais."
        ),
    )
    parser.add_argument(
        "--planilha",
        type=Path,
        default=DEFAULT_PRIORITY_PATH,
        help="Caminho da planilha com colunas Cod e Fornecedor.",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=OUTPUT_REPORT_PATH,
        help="Arquivo Excel de saida.",
    )
    parser.add_argument(
        "--atualizar-aba-cnpj",
        action="store_true",
        help="Cria ou atualiza a aba CNPJ_AUXILIAR na planilha original.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.planilha.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {args.planilha}")

    priority_items = iter_priority_rows(args.planilha)
    contract_records = load_contract_records()
    pdf_records = load_pdf_records()
    result_rows = build_result_rows(
        priority_items,
        contract_records,
        pdf_records,
    )
    write_report(result_rows, args.saida)
    if args.atualizar_aba_cnpj:
        write_pending_cnpj_sheet(args.planilha, result_rows)

    total_found = sum(
        1 for row in result_rows
        if row["status_match"] != "nao_encontrado"
    )
    print(f"Fornecedores na prioridade: {len(result_rows)}")
    print(f"Fornecedores com contrato ou PDF localizado: {total_found}")
    print(f"Relatorio gerado em: {args.saida}")
    print(f"PDFs copiados em: {MATCHED_PDF_DIR}")
    if args.atualizar_aba_cnpj:
        print(f"Aba CNPJ_AUXILIAR atualizada em: {args.planilha}")


if __name__ == "__main__":
    main()
