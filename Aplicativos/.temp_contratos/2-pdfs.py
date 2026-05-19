from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import shutil
import unicodedata
from collections import Counter
from functools import lru_cache
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageFilter, ImageOps
from rapidocr_onnxruntime import RapidOCR


BASE_DIR = Path(__file__).resolve().parent
PDF_DIR = BASE_DIR / "contratospdf"
OUTPUT_DIR = BASE_DIR / "saida"
SCHEMA_PATH = BASE_DIR / "schema_extracao.json"
REPORT_LOOKUP_PATH = OUTPUT_DIR / "contas_a_receber_resumo_fornecedor.json"
ABC_SUPPLIER_PATH = (
    BASE_DIR.parent / "pasta_abc_comprador" / "abc_comprador.csv"
)
SAMPLE_SIZE = 4
BATCH_SIZE = 25
CHECKPOINT_PATH = OUTPUT_DIR / "checkpoint_contratos.json"
INCREMENTAL_JSON_PATH = OUTPUT_DIR / "contratos_incremental.json"
INCREMENTAL_EXCEL_PATH = OUTPUT_DIR / "contratos_incremental.xlsx"
PRIORITY_PATH = BASE_DIR / "relatorio" / "lista_prioridade.xlsx"

SECTION_BOXES = {
    "cadastro_fornecedor": (0.18, 0.22, 0.83, 0.43),
    "condicoes_pagamento": (0.18, 0.41, 0.83, 0.58),
    "acordos_comerciais": (0.18, 0.57, 0.83, 0.74),
    "politica_trocas": (0.18, 0.73, 0.84, 0.84),
}

CORE_FIELDS = [
    "arquivo_pdf",
    "layout_tipo",
    "fornecedor_nome",
    "cnpj",
    "codigo_fornecedor_consinco",
    "fornecedor_descricao_atual",
    "a_receber",
    "dias_de_atraso",
    "data_assinatura",
    "ano_assinatura",
    "prazo_pagamento_dias",
    "forma_pagamento",
    "formas_pagamento_detectadas",
    "bonus_percentual",
    "cobranca_bonus",
    "forma_pagamento_bonus",
    "politica_trocas",
    "beneficios_detectados",
    "descontos_beneficios_detectados",
    "tipos_investimento_detectados",
    "observacoes_extracao",
]

OCR_ENGINE = None
MONTHS = {
    "janeiro": "01",
    "fevereiro": "02",
    "marco": "03",
    "abril": "04",
    "maio": "05",
    "junho": "06",
    "julho": "07",
    "agosto": "08",
    "setembro": "09",
    "outubro": "10",
    "novembro": "11",
    "dezembro": "12",
}
PAYMENT_PATTERNS = [
    ("BOLETO BANCARIO", ["boleto"]),
    ("DEPOSITO CONTA CORRENTE", ["deposit", "conta"]),
    ("DESCONTO EM DUPLICATAS", ["duplicat"]),
    ("BONIFICACAO EM PRODUTO", ["bonific", "produto"]),
    ("FACTORING", ["factoring"]),
    ("FOMENTO", ["fomento"]),
]
BENEFIT_PATTERNS = [
    ("BONUS_COMERCIAL", ["bonus"]),
    ("VERBA_ANIVERSARIO", ["aniversario"]),
    ("APOIO_LOGISTICO", ["apoio logistico"]),
    ("PROMOTOR", ["promotor"]),
    ("MIDIA_ESPACO_LOJA", ["midia", "espaco"]),
    ("BONIFICACAO_EM_PRODUTO", ["bonific", "produto"]),
    ("POLITICA_TROCAS", ["devol", "nota"]),
]
DISCOUNT_PATTERNS = [
    ("DESCONTO_BOLETO_BANCARIO", ["descont", "boleto"]),
    ("DESCONTO_EM_DUPLICATAS", ["duplicat"]),
    ("BOLETO_VALORES_PENDENTES", ["bloqueto", "pendent"]),
]
BONUS_PAYMENT_PATTERNS = [
    ("DESCONTO BOLETO BANCARIO", ["descont", "boleto"]),
    ("BONIFICACAO EM PRODUTO", ["bonific", "produto"]),
    ("DEPOSITO CONTA CORRENTE", ["deposit", "conta"]),
]
EXCHANGE_POLICY_PATTERNS = [
    ("DEVOLUCAO NOTA FISCAL", ["devol", "nota", "fisc"]),
    ("TROCA FISICA DE PRODUTO", ["troc", "fisic", "produt"]),
]
STRUCTURED_EXCHANGE_POLICIES = {
    "DEVOLUCAO NOTA FISCAL",
    "TROCA FISICA DE PRODUTO",
}
CHECKED_MARK_PATTERN = re.compile(r"[\(\[\{]\s*[xX]\s*[\)\]\}]")
UNCHECKED_MARK_PATTERN = re.compile(r"[\(\[\{]\s*[\)\]\}]")


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower()
    return re.sub(r"\s+", " ", ascii_text).strip()


def slugify(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", value)
    ascii_value = ascii_value.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.lower()
    ascii_value = re.sub(r"[^a-z0-9]+", "_", ascii_value).strip("_")
    return ascii_value or "arquivo"


def build_supplier_lookup_key(value: str) -> str:
    ascii_value = unicodedata.normalize("NFKD", str(value))
    ascii_value = ascii_value.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.upper()
    ascii_value = re.sub(
        (
            r"\b(LTDA|S/A|SA|ME|EPP|EIRELI|INDUSTRIA|COMERCIO|"
            r"DISTRIBUIDORA|DISTRIBUIDOR|ADM|DO|DE|DA|E)\b"
        ),
        " ",
        ascii_value,
    )
    ascii_value = re.sub(r"[^A-Z0-9]+", " ", ascii_value)
    return re.sub(r"\s+", " ", ascii_value).strip()


def is_supplier_match_acceptable(query: str, candidate: str) -> bool:
    if not query or not candidate:
        return False

    if query == candidate:
        return True

    if len(query) >= 8 and (query in candidate or candidate in query):
        return True

    query_tokens = {token for token in query.split() if len(token) >= 4}
    candidate_tokens = {
        token for token in candidate.split() if len(token) >= 4
    }
    overlap = query_tokens & candidate_tokens
    ratio = difflib.SequenceMatcher(None, query, candidate).ratio()

    if len(overlap) >= 2:
        return True
    return ratio >= 0.88


def normalize_cnpj_key(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if len(digits) == 14 else ""


def find_header_index(header: list[str], candidates: set[str]) -> int | None:
    for index, value in enumerate(header):
        if value in candidates:
            return index
    return None


@lru_cache(maxsize=1)
def load_priority_items() -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    if not PRIORITY_PATH.exists():
        return {}, {}, {}

    workbook = load_workbook(PRIORITY_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}, {}, {}

    header = [str(cell or "").strip().lower() for cell in rows[0]]
    code_index = find_header_index(header, {"cod", "codigo", "código"})
    supplier_index = find_header_index(header, {"fornecedor"})
    cnpj_index = next(
        (index for index, value in enumerate(header) if "cnpj" in value),
        None,
    )

    if supplier_index is None:
        return {}, {}, {}

    by_code: dict[str, int] = {}
    by_name: dict[str, int] = {}
    by_cnpj: dict[str, int] = {}

    for order, row in enumerate(rows[1:]):
        supplier_name = str(row[supplier_index] or "").strip()
        if not supplier_name:
            continue

        key_name = build_supplier_lookup_key(supplier_name)
        if key_name and key_name not in by_name:
            by_name[key_name] = order

        if code_index is not None:
            code_value = str(row[code_index] or "").strip()
            if code_value and code_value not in by_code:
                by_code[code_value] = order

        if cnpj_index is not None:
            cnpj_value = normalize_cnpj_key(row[cnpj_index])
            if cnpj_value and cnpj_value not in by_cnpj:
                by_cnpj[cnpj_value] = order

    return by_code, by_name, by_cnpj


def resolve_priority_order(
    supplier_name: object = "",
    supplier_code: object = "",
    supplier_cnpj: object = "",
) -> int:
    by_code, by_name, by_cnpj = load_priority_items()

    code_value = str(supplier_code or "").strip()
    if code_value and code_value in by_code:
        return by_code[code_value]

    cnpj_value = normalize_cnpj_key(supplier_cnpj)
    if cnpj_value and cnpj_value in by_cnpj:
        return by_cnpj[cnpj_value]

    name_key = build_supplier_lookup_key(str(supplier_name or ""))
    if name_key and name_key in by_name:
        return by_name[name_key]

    return 10**9


def sort_records_by_priority(
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    return sorted(
        records,
        key=lambda record: (
            0
            if resolve_priority_order(
                record.get("fornecedor_descricao_atual")
                or record.get("fornecedor_nome"),
                record.get("codigo_fornecedor_consinco"),
                record.get("cnpj"),
            )
            < 10**9
            else 1,
            resolve_priority_order(
                record.get("fornecedor_descricao_atual")
                or record.get("fornecedor_nome"),
                record.get("codigo_fornecedor_consinco"),
                record.get("cnpj"),
            ),
            str(record.get("fornecedor_descricao_atual") or ""),
            str(record.get("fornecedor_nome") or ""),
            str(record.get("arquivo_pdf") or ""),
        ),
    )


def sort_pdf_paths_by_priority(pdf_files: list[Path]) -> list[Path]:
    return sorted(
        pdf_files,
        key=lambda pdf_path: (
            0 if resolve_priority_order(pdf_path.stem) < 10**9 else 1,
            resolve_priority_order(pdf_path.stem),
            pdf_path.name,
        ),
    )


def safe_float(value: object) -> object:
    if value in (None, ""):
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def safe_int(value: object) -> object:
    if value in (None, ""):
        return ""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return ""


def load_report_supplier_lookup(
) -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]]]:
    if not REPORT_LOOKUP_PATH.exists():
        return {}, {}

    payload = json.loads(REPORT_LOOKUP_PATH.read_text(encoding="utf-8"))
    by_code: dict[str, dict[str, object]] = {}
    by_name: dict[str, dict[str, object]] = {}
    for item in payload:
        code = str(item.get("codigo_pessoa", "")).strip()
        name = str(item.get("fornecedor_relatorio_receber", "")).strip()
        normalized = build_supplier_lookup_key(name)
        normalized_item = {
            "codigo_fornecedor_consinco": code,
            "fornecedor_descricao_atual": name,
            "a_receber": safe_float(item.get("a_receber")),
            "dias_de_atraso": safe_int(item.get("dias_de_atraso")),
            "origem_fornecedor": "relatorio",
        }
        if code:
            by_code[code] = normalized_item
        if normalized:
            by_name[normalized] = normalized_item
    return by_code, by_name


def load_abc_supplier_lookup() -> dict[str, dict[str, object]]:
    if not ABC_SUPPLIER_PATH.exists():
        return {}

    lookup: dict[str, dict[str, object]] = {}
    with ABC_SUPPLIER_PATH.open(encoding="latin-1", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        for row in reader:
            code = str(
                row.get("CODIGO_FORNECEDOR_PRINCIPAL", ""),
            ).strip()
            description = str(
                row.get("FORNECEDOR_PRINCIPAL", ""),
            ).strip()
            if not code or code == "0" or not description:
                continue
            key = build_supplier_lookup_key(description)
            lookup.setdefault(
                key,
                {
                    "codigo_fornecedor_consinco": code,
                    "fornecedor_descricao_atual": description,
                    "origem_fornecedor": "abc_comprador",
                },
            )
    return lookup


def find_supplier_match(
    supplier_name: str,
    report_by_code: dict[str, dict[str, object]],
    report_by_name: dict[str, dict[str, object]],
    abc_by_name: dict[str, dict[str, object]],
) -> dict[str, object]:
    normalized = build_supplier_lookup_key(supplier_name)
    if not normalized:
        return {}

    if normalized in report_by_name:
        return dict(report_by_name[normalized])
    if normalized in abc_by_name:
        match = dict(abc_by_name[normalized])
        report_match = report_by_code.get(
            str(match.get("codigo_fornecedor_consinco", "")),
        )
        if report_match:
            match.update(report_match)
        return match

    report_candidates = difflib.get_close_matches(
        normalized,
        list(report_by_name.keys()),
        n=1,
        cutoff=0.72,
    )
    if report_candidates:
        candidate = report_candidates[0]
        if is_supplier_match_acceptable(normalized, candidate):
            return dict(report_by_name[candidate])

    abc_candidates = difflib.get_close_matches(
        normalized,
        list(abc_by_name.keys()),
        n=1,
        cutoff=0.74,
    )
    if abc_candidates:
        candidate = abc_candidates[0]
        if not is_supplier_match_acceptable(normalized, candidate):
            return {}
        match = dict(abc_by_name[candidate])
        report_match = report_by_code.get(
            str(match.get("codigo_fornecedor_consinco", "")),
        )
        if report_match:
            match.update(report_match)
        return match

    return {}


def enrich_supplier_fields(
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    report_by_code, report_by_name = load_report_supplier_lookup()
    abc_by_name = load_abc_supplier_lookup()

    grouped_records: dict[str, list[dict[str, object]]] = {}
    for record in records:
        cnpj_key = normalize_cnpj_key(record.get("cnpj"))
        if cnpj_key:
            group_key = f"CNPJ::{cnpj_key}"
        else:
            group_key = (
                "NOME::"
                f"{build_supplier_lookup_key(
                    record.get('fornecedor_nome', '')
                )}"
            )
        grouped_records.setdefault(group_key, []).append(dict(record))

    enriched_records: list[dict[str, object]] = []
    for group in grouped_records.values():
        representative = next(
            (
                item for item in group
                if item.get("fornecedor_nome")
            ),
            group[0],
        )
        match = find_supplier_match(
            str(representative.get("fornecedor_nome", "")),
            report_by_code,
            report_by_name,
            abc_by_name,
        )
        for item in group:
            enriched = dict(item)
            enriched.setdefault("codigo_fornecedor_consinco", "")
            enriched.setdefault("fornecedor_descricao_atual", "")
            enriched.setdefault("a_receber", "")
            enriched.setdefault("dias_de_atraso", "")
            if match:
                if not enriched.get("codigo_fornecedor_consinco"):
                    enriched["codigo_fornecedor_consinco"] = match.get(
                        "codigo_fornecedor_consinco",
                        "",
                    )
                if not enriched.get("fornecedor_descricao_atual"):
                    enriched["fornecedor_descricao_atual"] = match.get(
                        "fornecedor_descricao_atual",
                        "",
                    )
                if enriched.get("a_receber") in (None, ""):
                    enriched["a_receber"] = match.get("a_receber", "")
                if enriched.get("dias_de_atraso") in (None, ""):
                    enriched["dias_de_atraso"] = match.get(
                        "dias_de_atraso",
                        "",
                    )
            enriched_records.append(enriched)

    return enriched_records


def prepare_records_for_output(
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    finalized = [finalize_record(record) for record in records]
    enriched = enrich_supplier_fields(finalized)
    return sort_records_by_priority(enriched)


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def list_pdf_files(limit: int | None = None) -> list[Path]:
    pdf_files = sorted(
        path
        for path in PDF_DIR.glob("*.pdf")
        if path.is_file() and path.suffix.lower() == ".pdf"
    )
    pdf_files = sort_pdf_paths_by_priority(pdf_files)
    if limit is None:
        return pdf_files
    return pdf_files[:limit]


def load_schema() -> dict[str, object]:
    return json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))


def render_pdf_page(
    pdf_path: Path,
    page_index: int,
    output_path: Path,
    resolution: int = 180,
) -> Path:
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_index]
        image = page.to_image(resolution=resolution)
        image.save(output_path.as_posix(), format="PNG")
    return output_path


def preprocess_image(source_path: Path, target_path: Path) -> Path:
    image = Image.open(source_path).convert("L")
    image = ImageOps.autocontrast(image)
    image = image.point(lambda pixel: 255 if pixel > 180 else 0)
    image = image.filter(ImageFilter.MedianFilter(size=3))
    image.save(target_path)
    return target_path


def crop_sections(image_path: Path, target_dir: Path) -> dict[str, Path]:
    image = Image.open(image_path)
    width, height = image.size
    cropped_files: dict[str, Path] = {}

    for section_name, box in SECTION_BOXES.items():
        x0, y0, x1, y1 = box
        crop_box = (
            int(width * x0),
            int(height * y0),
            int(width * x1),
            int(height * y1),
        )
        cropped = image.crop(crop_box)
        output_path = target_dir / f"{section_name}.png"
        cropped.save(output_path)
        cropped_files[section_name] = output_path

    return cropped_files


def ocr_lines(image_path: Path) -> list[str]:
    global OCR_ENGINE
    if OCR_ENGINE is None:
        OCR_ENGINE = RapidOCR()
    result, _ = OCR_ENGINE(image_path.as_posix())
    if not result:
        return []
    return [item[1].strip() for item in result if item and item[1].strip()]


def extract_pdf_text_lines(
    pdf_path: Path,
    page_index: int,
) -> list[str]:
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_index]
        text = page.extract_text() or ""
    lines = [clean_value(line) for line in text.splitlines() if line.strip()]
    return [line for line in lines if line]


def clean_value(raw_value: str) -> str:
    value = raw_value.replace("_", " ").strip(" .:-")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_digits(value: str) -> str:
    converted = value.lower().translate(
        str.maketrans({
            "o": "0",
            "q": "0",
            "g": "6",
            "s": "5",
            "b": "8",
            "i": "1",
            "l": "1",
        })
    )
    return re.sub(r"\D", "", converted)


def detect_layout_type(full_page_lines: list[str]) -> str:
    full_text = " ".join(normalize_text(line) for line in full_page_lines)
    if "acordo comercial 2013" in full_text:
        return "ACORDO_COMERCIAL_2013"
    if "acordo comercial" in full_text:
        return "ACORDO_COMERCIAL"
    return "NAO_IDENTIFICADO"


def extract_supplier_name(lines: list[str], fallback_name: str) -> str:
    for line in lines:
        normalized = normalize_text(line)
        if "fornecedor" in normalized:
            value = re.split(
                r"fornecedor\s*",
                line,
                flags=re.IGNORECASE,
                maxsplit=1,
            )[-1]
            value = re.sub(r"^[\s:.-]+", "", value)
            value = clean_value(value)
            value = re.sub(r"\b(EPP|LTDA|ME)\b$", "", value).strip()
            invalid_tokens = [
                "concorda",
                "percentual",
                "acordo",
                "compradora",
                "supermercado",
            ]
            is_invalid = any(
                token in normalize_text(value) for token in invalid_tokens
            )
            if value and not is_invalid:
                return value
    return fallback_name


def extract_cnpj(lines: list[str]) -> str:
    pattern = re.compile(r"\d{2}[\.,]?\d{3}[\.,]?\d{3}/\d{4}-\d{2}")
    for line in lines:
        match = pattern.search(line)
        if match:
            digits = re.sub(r"\D", "", match.group(0))
            if len(digits) == 14:
                return (
                    f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/"
                    f"{digits[8:12]}-{digits[12:14]}"
                )
    return ""


def extract_payment_days(lines: list[str]) -> int | None:
    for index, line in enumerate(lines):
        normalized = normalize_text(line)
        candidates = [line]
        if "prazo" in normalized and index + 1 < len(lines):
            candidates.append(lines[index + 1])
        for candidate in candidates:
            match = re.search(r"(\d{1,3})\s*dias", normalize_text(candidate))
            if match:
                return int(match.group(1))
    return None


def extract_main_payment_method(lines: list[str]) -> str:
    detected = detect_payment_types(lines)
    for payment_type in [
        "BOLETO BANCARIO",
        "DEPOSITO CONTA CORRENTE",
        "DESCONTO EM DUPLICATAS",
        "BONIFICACAO EM PRODUTO",
    ]:
        if payment_type in detected:
            return payment_type
    return ""


def extract_bonus_percentage(lines: list[str]) -> float | None:
    for line in lines:
        normalized = normalize_text(line)
        if "bonus" in normalized or "percen" in normalized or "%" in line:
            match = re.search(r"(\d{1,2}(?:[\.,]\d{1,2})?)\)?\s*%", line)
            if match:
                return float(match.group(1).replace(",", "."))
    return None


def extract_bonus_charge(lines: list[str]) -> str:
    for line in lines:
        normalized = normalize_text(line)
        if "mens" in normalized or "ens" in normalized:
            return "MENSAL"
        if "trimes" in normalized:
            return "TRIMESTRAL"
        if "semes" in normalized:
            return "SEMESTRAL"
    return ""


def extract_bonus_payment_method(lines: list[str]) -> str:
    detected = detect_labeled_options(lines, BONUS_PAYMENT_PATTERNS)
    for payment_type in [
        "DESCONTO BOLETO BANCARIO",
        "BONIFICACAO EM PRODUTO",
        "DEPOSITO CONTA CORRENTE",
    ]:
        if payment_type in detected:
            return payment_type
    return ""


def extract_exchange_policy(lines: list[str]) -> str:
    detected = detect_labeled_options(lines, EXCHANGE_POLICY_PATTERNS)
    if len(detected) == 1:
        return detected[0]

    normalized_block = " ".join(normalize_text(line) for line in lines)
    has_devolucao = all(
        token in normalized_block for token in ["devol", "nota", "fisc"]
    )
    has_troca = all(
        token in normalized_block for token in ["troc", "fisic", "produt"]
    )

    if has_devolucao and not has_troca:
        return "DEVOLUCAO NOTA FISCAL"
    if has_troca and not has_devolucao:
        return "TROCA FISICA DE PRODUTO"
    return ""


def sanitize_exchange_policy(value: object) -> str:
    if value in (None, ""):
        return ""

    text = clean_value(str(value))
    if not text:
        return ""

    if text in STRUCTURED_EXCHANGE_POLICIES:
        return text

    segments = [
        clean_value(part)
        for part in re.split(r"[\r\n|]+", text)
        if clean_value(part)
    ]
    if not segments:
        return ""

    detected = extract_exchange_policy(segments)
    if detected:
        return detected

    return ""


def has_structured_exchange_policy(value: object) -> bool:
    if not isinstance(value, str):
        return False
    return value in STRUCTURED_EXCHANGE_POLICIES


def split_multi_value_field(value: object) -> list[str]:
    if not isinstance(value, str):
        return []
    items = [item.strip() for item in value.split(";")]
    return [item for item in items if item]


def has_checkbox_syntax(value: str) -> bool:
    return bool(
        CHECKED_MARK_PATTERN.search(value)
        or UNCHECKED_MARK_PATTERN.search(value)
    )


def tokens_in_text(value: str, tokens: list[str]) -> bool:
    return all(token in value for token in tokens)


def has_marked_option(value: str, tokens: list[str]) -> bool:
    if not tokens_in_text(value, tokens):
        return False
    spans: list[tuple[int, int]] = []
    for token in tokens:
        start = value.find(token)
        if start < 0:
            continue
        spans.append((start, start + len(token)))
    if not spans:
        return False
    option_start = min(start for start, _ in spans)
    option_end = max(end for _, end in spans)
    before_segment = value[max(option_start - 12, 0):option_start]
    after_segment = value[option_end:min(option_end + 12, len(value))]
    after_match = re.match(
        r"^\s*[\(\[\{]\s*([xX]?)\s*[\)\]\}]",
        after_segment,
    )
    if after_match:
        return bool(after_match.group(1))
    before_match = re.search(
        r"[\(\[\{]\s*([xX]?)\s*[\)\]\}]\s*$",
        before_segment,
    )
    if before_match:
        return bool(before_match.group(1))
    return False


def detect_labeled_options(
    lines: list[str],
    patterns: list[tuple[str, list[str]]],
) -> list[str]:
    normalized_lines = [normalize_text(line) for line in lines if line.strip()]
    normalized_block = " ".join(normalized_lines)

    if any(has_checkbox_syntax(line) for line in normalized_lines):
        detected: list[str] = []
        for label, tokens in patterns:
            if any(
                has_marked_option(line, tokens)
                for line in normalized_lines
            ):
                detected.append(label)
        return detected

    detected = []
    for label, tokens in patterns:
        if tokens_in_text(normalized_block, tokens):
            detected.append(label)
    return detected


def build_benefit_columns(record: dict[str, object]) -> dict[str, object]:
    columns: dict[str, object] = {}
    benefits = split_multi_value_field(
        record.get("beneficios_detectados"),
    )
    for benefit in benefits:
        columns[f"beneficio_{slugify(benefit)}"] = "SIM"
    return columns


def build_discount_columns(record: dict[str, object]) -> dict[str, object]:
    columns: dict[str, object] = {}
    discounts = split_multi_value_field(
        record.get("descontos_beneficios_detectados"),
    )
    for discount in discounts:
        columns[f"desconto_{slugify(discount)}"] = "SIM"
    return columns


def detect_payment_types(lines: list[str]) -> list[str]:
    return detect_labeled_options(lines, PAYMENT_PATTERNS)


def detect_benefits(lines: list[str]) -> list[str]:
    return detect_labeled_options(lines, BENEFIT_PATTERNS)


def detect_discount_benefits(lines: list[str]) -> list[str]:
    return detect_labeled_options(lines, DISCOUNT_PATTERNS)


def normalize_month_token(token: str) -> str | None:
    cleaned = re.sub(r"[^a-z]", "", normalize_text(token))
    if not cleaned:
        return None
    if cleaned in MONTHS:
        return cleaned
    close = difflib.get_close_matches(cleaned, MONTHS.keys(), n=1, cutoff=0.6)
    if close:
        return close[0]
    return None


def extract_signature_date(lines: list[str]) -> tuple[str, str]:
    joined = " ".join(normalize_text(line) for line in lines)
    date_pattern = re.compile(
        r"(\d{1,2})\s*de\s*([a-z]+)\s*de\s*([0-9a-z]{2,6})"
    )
    for match in date_pattern.finditer(joined):
        day = int(match.group(1))
        month_token = normalize_month_token(match.group(2))
        year_digits = normalize_digits(match.group(3))
        if len(year_digits) == 2:
            year_digits = f"20{year_digits}"
        if month_token and len(year_digits) == 4 and 1 <= day <= 31:
            formatted = f"{day:02d}/{MONTHS[month_token]}/{year_digits}"
            return formatted, year_digits

    numeric_pattern = re.compile(
        r"(\d{1,2})\s*[/-]\s*(\d{1,2})\s*[/-]\s*(\d{2,4})"
    )
    for match in numeric_pattern.finditer(joined):
        day = int(match.group(1))
        month = int(match.group(2))
        year_digits = normalize_digits(match.group(3))
        if len(year_digits) == 2:
            year_digits = f"20{year_digits}"
        if len(year_digits) == 4 and 1 <= day <= 31 and 1 <= month <= 12:
            return f"{day:02d}/{month:02d}/{year_digits}", year_digits

    return "", ""


def build_investment_columns(record: dict[str, object]) -> dict[str, object]:
    investment_types: list[str] = []
    has_bonus = any(
        record.get(key) not in (None, "")
        for key in [
            "bonus_percentual",
            "cobranca_bonus",
            "forma_pagamento_bonus",
        ]
    )
    has_exchange_policy = has_structured_exchange_policy(
        record.get("politica_trocas"),
    )
    if has_bonus:
        investment_types.append("BONUS_COMERCIAL")
    if has_exchange_policy:
        investment_types.append("POLITICA_TROCAS")

    return {
        "investimento_bonus_comercial": "SIM" if has_bonus else "NAO",
        "investimento_bonus_percentual": record.get("bonus_percentual"),
        "investimento_bonus_cobranca": record.get("cobranca_bonus", ""),
        "investimento_bonus_forma_pagamento": record.get(
            "forma_pagamento_bonus",
            "",
        ),
        "investimento_politica_trocas": (
            record.get("politica_trocas", "") if has_exchange_policy else ""
        ),
        "tipos_investimento_detectados": "; ".join(investment_types),
    }


def finalize_record(record: dict[str, object]) -> dict[str, object]:
    normalized = dict(record)
    normalized.setdefault("codigo_fornecedor_consinco", "")
    normalized.setdefault("fornecedor_descricao_atual", "")
    normalized.setdefault("a_receber", "")
    normalized.setdefault("dias_de_atraso", "")
    normalized["politica_trocas"] = sanitize_exchange_policy(
        normalized.get("politica_trocas"),
    )
    normalized.update(build_investment_columns(normalized))
    normalized.update(build_benefit_columns(normalized))
    normalized.update(build_discount_columns(normalized))
    return normalized


def get_pdf_page_count(pdf_path: Path) -> int:
    with pdfplumber.open(pdf_path) as pdf:
        return len(pdf.pages)


def process_page(
    pdf_path: Path,
    page_index: int,
    work_dir: Path,
    prefix: str,
) -> tuple[Path, list[str]]:
    render_path = work_dir / f"{prefix}_render.png"
    processed_path = work_dir / f"{prefix}_preprocessada.png"
    render_pdf_page(pdf_path, page_index, render_path)
    preprocess_image(render_path, processed_path)
    return processed_path, ocr_lines(processed_path)


def extract_contract_data(
    pdf_path: Path,
    keep_debug_assets: bool = False,
    scan_last_page: bool = True,
) -> tuple[dict[str, object], dict[str, list[str]]]:
    ensure_output_dir()
    base_work_dir = OUTPUT_DIR / ("debug" if keep_debug_assets else "_tmp")
    work_dir = base_work_dir / slugify(pdf_path.stem)
    if work_dir.exists():
        shutil.rmtree(work_dir)
    work_dir.mkdir(parents=True, exist_ok=True)

    page_count = get_pdf_page_count(pdf_path)
    last_page_index = max(page_count - 1, 0)
    first_page_lines = extract_pdf_text_lines(pdf_path, 0)
    last_page_lines = []
    if scan_last_page:
        last_page_lines = extract_pdf_text_lines(pdf_path, last_page_index)

    uses_native_text = len(first_page_lines) >= 8
    section_ocr: dict[str, list[str]]

    if uses_native_text:
        section_ocr = {
            name: list(first_page_lines)
            for name in SECTION_BOXES
        }
    else:
        first_page_path, first_page_lines = process_page(
            pdf_path,
            0,
            work_dir,
            "pagina_1",
        )
        if scan_last_page:
            _, last_page_lines = process_page(
                pdf_path,
                last_page_index,
                work_dir,
                "pagina_final",
            )
        section_ocr = {
            name: list(first_page_lines)
            for name in SECTION_BOXES
        }
        if len(first_page_lines) < 8:
            section_images = crop_sections(first_page_path, work_dir)
            section_ocr = {
                name: ocr_lines(path) for name, path in section_images.items()
            }
    section_ocr["pagina_1_completa"] = first_page_lines
    section_ocr["pagina_final_completa"] = last_page_lines

    fallback_name = pdf_path.stem
    supplier_lines = section_ocr["cadastro_fornecedor"]
    payment_lines = section_ocr["condicoes_pagamento"] + supplier_lines
    agreement_lines = section_ocr["acordos_comerciais"] + payment_lines
    exchange_lines = section_ocr["politica_trocas"] + agreement_lines
    all_relevant_lines = first_page_lines + last_page_lines + agreement_lines
    signature_source = last_page_lines or first_page_lines
    signature_date, signature_year = extract_signature_date(signature_source)
    payment_types = detect_payment_types(all_relevant_lines)
    benefits = detect_benefits(all_relevant_lines)
    discount_benefits = detect_discount_benefits(all_relevant_lines)

    extracted = {
        "arquivo_pdf": pdf_path.name,
        "layout_tipo": detect_layout_type(first_page_lines),
        "fornecedor_nome": extract_supplier_name(
            supplier_lines,
            fallback_name,
        ),
        "linhas_extraidas_pagina_1": len(first_page_lines),
        "linhas_extraidas_pagina_final": len(last_page_lines),
        "cnpj": extract_cnpj(supplier_lines),
        "data_assinatura": signature_date,
        "ano_assinatura": signature_year,
        "prazo_pagamento_dias": extract_payment_days(payment_lines),
        "forma_pagamento": extract_main_payment_method(all_relevant_lines),
        "formas_pagamento_detectadas": "; ".join(payment_types),
        "bonus_percentual": extract_bonus_percentage(agreement_lines),
        "cobranca_bonus": extract_bonus_charge(agreement_lines),
        "forma_pagamento_bonus": extract_bonus_payment_method(agreement_lines),
        "politica_trocas": extract_exchange_policy(exchange_lines),
        "beneficios_detectados": "; ".join(benefits),
        "descontos_beneficios_detectados": "; ".join(discount_benefits),
        "observacoes_extracao": (
            "Texto nativo do PDF quando disponivel; caso contrario OCR da "
            "primeira e da ultima pagina. Revisar contratos com layout "
            "diferente, paginas extras relevantes ou campos vazios."
        ),
    }
    extracted = finalize_record(extracted)

    if not keep_debug_assets:
        shutil.rmtree(work_dir, ignore_errors=True)

    return extracted, section_ocr


def analyze_contracts(
    pdf_files: list[Path],
    keep_debug_assets: bool,
) -> tuple[list[dict[str, object]], dict[str, dict[str, list[str]]]]:
    records: list[dict[str, object]] = []
    debug_payload: dict[str, dict[str, list[str]]] = {}

    for pdf_path in pdf_files:
        record, section_ocr = extract_contract_data(
            pdf_path,
            keep_debug_assets=keep_debug_assets,
        )
        records.append(record)
        if keep_debug_assets:
            debug_payload[pdf_path.name] = section_ocr

    return records, debug_payload


def load_checkpoint() -> dict[str, object]:
    if not CHECKPOINT_PATH.exists():
        return {"records": [], "processed_files": []}
    return json.loads(CHECKPOINT_PATH.read_text(encoding="utf-8"))


def save_checkpoint(records: list[dict[str, object]]) -> None:
    payload = {
        "records": records,
        "processed_files": [record["arquivo_pdf"] for record in records],
        "total_processados": len(records),
    }
    CHECKPOINT_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def analyze_contracts_in_batches(
    pdf_files: list[Path],
    batch_size: int = BATCH_SIZE,
) -> list[dict[str, object]]:
    checkpoint = load_checkpoint()
    records = list(checkpoint.get("records", []))
    processed_files = set(checkpoint.get("processed_files", []))
    pending_files = [
        pdf_path
        for pdf_path in pdf_files
        if pdf_path.name not in processed_files
    ]

    if records:
        print(
            f"Retomando checkpoint com {len(records)} contratos processados."
        )

    for start in range(0, len(pending_files), batch_size):
        batch = pending_files[start:start + batch_size]
        for pdf_path in batch:
            record, _ = extract_contract_data(
                pdf_path,
                keep_debug_assets=False,
            )
            records.append(record)
        save_checkpoint(records)
        save_consolidated_outputs(records)
        end = start + len(batch)
        message = (
            f"Lote concluido: {end}/{len(pending_files)} "
            "pendentes desta rodada."
        )
        print(message)

    return records


def summarize_sample(
    records: list[dict[str, object]],
) -> list[dict[str, object]]:
    summary: list[dict[str, object]] = []
    critical_fields = [
        "fornecedor_nome",
        "cnpj",
        "ano_assinatura",
        "prazo_pagamento_dias",
        "forma_pagamento",
        "bonus_percentual",
        "politica_trocas",
    ]

    for record in records:
        filled = [
            field
            for field in critical_fields
            if record.get(field) not in (None, "")
        ]
        missing = [field for field in critical_fields if field not in filled]
        summary.append(
            {
                "arquivo_pdf": record["arquivo_pdf"],
                "layout_tipo": record["layout_tipo"],
                "campos_preenchidos": len(filled),
                "campos_vazios": ", ".join(missing),
                "tipos_investimento_detectados": record[
                    "tipos_investimento_detectados"
                ],
            }
        )
    return summary


def save_sample_outputs(
    sample_records: list[dict[str, object]],
    sample_summary: list[dict[str, object]],
    sample_debug: dict[str, dict[str, list[str]]],
) -> None:
    sample_records = prepare_records_for_output(sample_records)
    sample_json = OUTPUT_DIR / "amostra_contratos.json"
    sample_debug_json = OUTPUT_DIR / "amostra_contratos_debug.json"
    sample_excel = OUTPUT_DIR / "amostra_contratos.xlsx"

    sample_json.write_text(
        json.dumps(sample_records, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    sample_debug_json.write_text(
        json.dumps(sample_debug, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_workbook(
        sample_excel,
        [
            (
                "amostra",
                sample_records,
                order_columns_from_records(sample_records),
            ),
            (
                "resumo",
                sample_summary,
                order_columns_from_records(sample_summary),
            ),
        ],
    )


def order_columns_from_records(records: list[dict[str, object]]) -> list[str]:
    if not records:
        return []
    present_columns = {
        key
        for record in records
        for key in record.keys()
    }
    ordered_columns = [
        column for column in CORE_FIELDS if column in present_columns
    ]
    dynamic_columns = sorted(
        column for column in present_columns if column not in ordered_columns
    )
    return ordered_columns + dynamic_columns


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    records: list[dict[str, object]],
    columns: list[str],
) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    if not columns:
        worksheet.append(["sem_dados"])
        return
    worksheet.append(columns)
    for record in records:
        worksheet.append([record.get(column, "") for column in columns])


def write_workbook(
    output_path: Path,
    sheets: list[tuple[str, list[dict[str, object]], list[str]]],
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, records, columns in sheets:
        write_sheet(workbook, sheet_name, records, columns)
    workbook.save(output_path)


def save_consolidated_outputs(all_records: list[dict[str, object]]) -> None:
    all_records = prepare_records_for_output(all_records)
    consolidated_json = OUTPUT_DIR / "contratos_consolidados.json"
    consolidated_excel = OUTPUT_DIR / "contratos_consolidados.xlsx"
    payment_summary_json = OUTPUT_DIR / "resumo_formas_pagamento.json"
    benefit_summary_json = OUTPUT_DIR / "resumo_beneficios.json"

    consolidated_json.write_text(
        json.dumps(all_records, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    ordered_columns = order_columns_from_records(all_records)

    payment_counter = Counter()
    benefit_counter = Counter()
    for record in all_records:
        payment_items = str(
            record.get("formas_pagamento_detectadas", ""),
        ).split("; ")
        benefit_items = str(
            record.get("beneficios_detectados", ""),
        ).split("; ")
        for item in payment_items:
            if item:
                payment_counter[item] += 1
        for item in benefit_items:
            if item:
                benefit_counter[item] += 1

    payment_summary = [
        {
            "forma_pagamento_detectada": label,
            "quantidade": count,
        }
        for label, count in payment_counter.most_common()
    ]
    benefit_summary = [
        {
            "beneficio_detectado": label,
            "quantidade": count,
        }
        for label, count in benefit_counter.most_common()
    ]

    write_workbook(
        consolidated_excel,
        [
            ("contratos", all_records, ordered_columns),
            (
                "resumo_pagamentos",
                payment_summary,
                ["forma_pagamento_detectada", "quantidade"],
            ),
            (
                "resumo_beneficios",
                benefit_summary,
                ["beneficio_detectado", "quantidade"],
            ),
        ],
    )

    payment_summary_json.write_text(
        json.dumps(payment_summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    benefit_summary_json.write_text(
        json.dumps(benefit_summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def load_incremental_records() -> list[dict[str, object]]:
    if INCREMENTAL_JSON_PATH.exists():
        records = json.loads(INCREMENTAL_JSON_PATH.read_text(encoding="utf-8"))
        return [finalize_record(record) for record in records]

    first_contract_path = OUTPUT_DIR / "primeiro_contrato_extraido.json"
    if first_contract_path.exists():
        first_content = first_contract_path.read_text(encoding="utf-8")
        first_record = json.loads(first_content)
        normalized_first = dict(first_record)
        for missing_key in [
            "layout_tipo",
            "data_assinatura",
            "ano_assinatura",
            "formas_pagamento_detectadas",
            "beneficios_detectados",
            "descontos_beneficios_detectados",
            "tipos_investimento_detectados",
        ]:
            normalized_first.setdefault(missing_key, "")
        return [finalize_record(normalized_first)]

    pdf_files = list_pdf_files(limit=1)
    if not pdf_files:
        return []
    first_record, _ = extract_contract_data(
        pdf_files[0],
        keep_debug_assets=True,
    )
    return [first_record]


def save_incremental_outputs(records: list[dict[str, object]]) -> None:
    records = prepare_records_for_output(records)
    INCREMENTAL_JSON_PATH.write_text(
        json.dumps(records, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    write_workbook(
        INCREMENTAL_EXCEL_PATH,
        [("contratos", records, order_columns_from_records(records))],
    )


def process_next_contract_incremental() -> list[dict[str, object]]:
    records = load_incremental_records()
    processed = {record.get("arquivo_pdf", "") for record in records}
    pdf_files = list_pdf_files()
    next_pdf = next(
        (pdf_path for pdf_path in pdf_files if pdf_path.name not in processed),
        None,
    )

    if next_pdf is None:
        save_incremental_outputs(records)
        return records

    next_record, _ = extract_contract_data(next_pdf, keep_debug_assets=True)
    records.append(next_record)
    save_incremental_outputs(records)
    return records


def print_stage_header(title: str) -> None:
    print(f"\n{'=' * 12} {title} {'=' * 12}")


def run_full_pipeline() -> None:
    ensure_output_dir()
    schema = load_schema()
    pdf_files = list_pdf_files()
    if not pdf_files:
        raise FileNotFoundError(f"Nenhum PDF encontrado em {PDF_DIR}")

    print_stage_header("ETAPA 1 - AMOSTRA")
    sample_files = pdf_files[:SAMPLE_SIZE]
    sample_records, sample_debug = analyze_contracts(
        sample_files,
        keep_debug_assets=True,
    )
    sample_summary = summarize_sample(sample_records)
    save_sample_outputs(sample_records, sample_summary, sample_debug)
    print(f"Contratos avaliados na amostra: {len(sample_records)}")
    print(json.dumps(sample_summary, indent=2, ensure_ascii=False))

    print_stage_header("ETAPA 2 - COLUNAS FINAIS")
    final_columns = schema.get("colunas_excel", [])
    print("Colunas definidas para consolidacao:")
    for column in final_columns:
        print(f"- {column}")

    print_stage_header("ETAPA 3 - CONSOLIDACAO")
    all_records = analyze_contracts_in_batches(pdf_files)
    save_consolidated_outputs(all_records)
    print(f"Total de contratos consolidados: {len(all_records)}")
    print(f"Arquivos gerados em: {OUTPUT_DIR}")


def run_incremental_next() -> None:
    ensure_output_dir()
    records = process_next_contract_incremental()
    latest = records[-1] if records else {}
    print_stage_header("MODO INCREMENTAL")
    print(f"Contratos na base incremental: {len(records)}")
    if latest:
        print(f"Ultimo contrato adicionado: {latest.get('arquivo_pdf', '')}")
        print(json.dumps(latest, indent=2, ensure_ascii=False))
    print(f"Arquivos gerados em: {OUTPUT_DIR}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extrator incremental de contratos",
    )
    parser.add_argument(
        "--modo",
        choices=["completo", "incremental"],
        default="incremental",
        help=(
            "Executa a consolidacao completa ou adiciona somente o "
            "proximo contrato."
        ),
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.modo == "completo":
        run_full_pipeline()
        return
    run_incremental_next()


if __name__ == "__main__":
    main()

