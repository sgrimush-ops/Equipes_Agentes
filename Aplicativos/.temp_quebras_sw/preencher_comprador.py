"""
Preenche a coluna 'Apelido Comprador' no itens_extraidos_preenchido.xlsx
usando o mapeamento CODIGO_PRODUTO → COMPRADOR do query.parquet.
"""
import os
from pathlib import Path
import pandas as pd
import openpyxl


def carregar_mapa_comprador() -> dict:
    """Carrega o query.parquet e retorna dict {codigo_produto_int: apelido_comprador}."""
    caminho = Path(__file__).parent.parent / "import_querys" / "query.parquet"
    if not caminho.exists():
        raise FileNotFoundError(f"query.parquet não encontrado: {caminho}")

    df = pd.read_parquet(caminho, columns=["CODIGO_PRODUTO", "COMPRADOR"])
    mapa = df.drop_duplicates(subset="CODIGO_PRODUTO", keep="first")
    resultado = dict(zip(mapa["CODIGO_PRODUTO"].astype(int), mapa["COMPRADOR"]))
    print(f"[MAPA] {len(resultado)} produtos com comprador mapeado")
    return resultado


def preencher_comprador() -> None:
    """Lê o Excel, cruza com o mapa e adiciona a coluna Apelido Comprador via openpyxl."""
    caminho_excel = Path("itens_extraidos_preenchido.xlsx")
    if not caminho_excel.exists():
        raise FileNotFoundError(f"Excel não encontrado: {caminho_excel}")

    mapa = carregar_mapa_comprador()

    print(f"\n[LENDO] {caminho_excel}...")
    df = pd.read_excel(caminho_excel, engine="openpyxl")
    print(f"[LENDO] {len(df):,} linhas carregadas")

    col_cod = "Codigo Consico"
    codigos_validos = df[col_cod].dropna()
    codigos_int = codigos_validos.apply(lambda x: int(float(x)))
    total_com_codigo = len(codigos_int)

    preenchidos = 0
    nao_encontrados = set()
    comprador_col = []

    for _, row in df.iterrows():
        cod = row[col_cod]
        if pd.isna(cod):
            comprador_col.append("")
            continue
        try:
            cod_int = int(float(cod))
        except (ValueError, TypeError):
            comprador_col.append("")
            continue

        apelido = mapa.get(cod_int)
        if apelido:
            comprador_col.append(str(apelido))
            preenchidos += 1
        else:
            comprador_col.append("")
            nao_encontrados.add(cod_int)

    print(f"\n[RESULTADO]")
    print(f"  Linhas com código Consico: {total_com_codigo:,}")
    print(f"  Preenchidas com comprador: {preenchidos:,}")
    print(f"  Produtos sem comprador no mapa: {len(nao_encontrados):,}")

    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active

    ultima_col = ws.max_column + 1
    ws.cell(row=1, column=ultima_col, value="Apelido Comprador")

    for i, apelido in enumerate(comprador_col):
        if apelido:
            ws.cell(row=i + 2, column=ultima_col, value=apelido)

    wb.save(caminho_excel)
    wb.close()

    print(f"\n[SALVO] Coluna 'Apelido Comprador' adicionada na coluna {ultima_col}")
    print(f"[SALVO] Arquivo: {caminho_excel.resolve()}")


if __name__ == "__main__":
    os.chdir(Path(__file__).parent.resolve())
    print("=" * 60)
    print("  PREENCHIMENTO DE COMPRADOR — itens_extraidos_preenchido.xlsx")
    print("=" * 60)
    preencher_comprador()
    print("\n" + "=" * 60)
    print("  CONCLUÍDO")
    print("=" * 60)
